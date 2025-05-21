import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os



def validate_inputs():
    name = entry_name.get()
    course = entry_course.get()
    grade = entry_grade.get()

    if not name or not course or not grade:
        messagebox.showerror("Input Error", "All fields are required!")
        return False
    return True

def save_to_excel():
    if not validate_inputs():
        return

    name = entry_name.get()
    course = entry_course.get()
    grade = entry_grade.get()

    filename = "gradessssssssssssssssss.xlsx"

    # Check if Excel file exists, if not create it with header
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Name", "Course", "Grade"])  # Header row
    else:
        wb = load_workbook(filename)
        if "Grades" not in wb.sheetnames:
            ws = wb.create_sheet("Grades")
            ws.append(["Name", "Course", "Grade"])  # Header row
        else:
            ws = wb["Grades"]

    ws.append([name, course, grade])
    wb.save(filename)

    messagebox.showinfo("Done", "Your entry has been recorded!")

    entry_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_grade.delete(0, tk.END)

def show_data():
    filename = "gradessssssssssssssssss.xlsx"

    if not os.path.exists(filename):
        messagebox.showerror("Error", "No data file found!")
        return

    wb = load_workbook(filename)

    if "Grades" not in wb.sheetnames:
        messagebox.showerror("Error", "No 'Grades' worksheet found!")
        return

    ws = wb["Grades"]

    data = tk.Toplevel(main_window)
    data.title("Student Data")
    data.geometry("300x200")
    data.configure(bg="light gray")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = tk.Label(data, text=value, bg="light gray", padx=4, pady=2)
            label.grid(row=i, column=j)

# --------------------- Main Window Setup ---------------------

main_window = tk.Tk()
main_window.geometry("260x250")
main_window.title("Grade Report")
main_window.configure(bg="light gray")

# --------------------- Header ---------------------

label_header = tk.Label(main_window, text="Grade Report", font=("arial", 14), bg="light gray")
label_header.pack(pady=10)

# --------------------- Frame for Form ---------------------

form_frame = tk.Frame(main_window, bg="light gray")
form_frame.pack()

# --------------------- Labels ---------------------

label_name = tk.Label(form_frame, text="Name", font=("arial", 10), bg="light gray")
label_name.grid(row=0, column=0, sticky="w", padx=3, pady=3)

label_course = tk.Label(form_frame, text="Course", font=("arial", 10), bg="light gray")
label_course.grid(row=1, column=0, sticky="w", padx=3, pady=3)

label_grade = tk.Label(form_frame, text="Grade", font=("arial", 10), bg="light gray")
label_grade.grid(row=2, column=0, sticky="w", padx=3, pady=3)

# --------------------- Entry Fields ---------------------

entry_name = tk.Entry(form_frame, width=18)
entry_name.grid(row=0, column=1, padx=3, pady=3)

entry_course = tk.Entry(form_frame, width=18)
entry_course.grid(row=1, column=1, padx=3, pady=3)

entry_grade = tk.Entry(form_frame, width=18)
entry_grade.grid(row=2, column=1, padx=3, pady=3)

# --------------------- Buttons ---------------------

button_save = tk.Button(form_frame, text="Save", font=("arial", 9), width=10, command=save_to_excel)
button_save.grid(row=3, column=0, padx=4, pady=10)

button_view = tk.Button(form_frame, text="View Data", font=("arial", 9), width=10, command=show_data)
button_view.grid(row=3, column=1, padx=4, pady=10)

# --------------------- Run Application ---------------------

main_window.mainloop()
